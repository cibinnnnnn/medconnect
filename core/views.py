from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from accounts.models import Doctor, Patient, User
from appointments.models import Appointment
from django.db.models import Count
import csv
import io
from datetime import datetime
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def index(request):
    """Homepage view"""
    departments = [
        {'name': 'Cardiology', 'icon': '❤️', 'description': 'Heart & Cardiovascular Care'},
        {'name': 'Neurology', 'icon': '🧠', 'description': 'Brain & Nervous System'},
        {'name': 'Orthopedics', 'icon': '🦴', 'description': 'Bone & Joint Care'},
        {'name': 'Pediatrics', 'icon': '👶', 'description': 'Child Healthcare'},
        {'name': 'Dermatology', 'icon': '✨', 'description': 'Skin Care'},
        {'name': 'General Medicine', 'icon': '🏥', 'description': 'General Health'},
    ]
    
    testimonials = [
        {
            'name': 'Sarah Johnson',
            'rating': 5,
            'text': 'MedConnect made booking appointments so easy! The AI symptom checker helped me understand my condition better.',
        },
        {
            'name': 'Michael Chen',
            'rating': 5,
            'text': 'Great platform! I can access all my medical records in one place. Highly recommended!',
        },
        {
            'name': 'Emily Davis',
            'rating': 4,
            'text': 'Very convenient and user-friendly. The QR code feature is innovative.',
        },
    ]
    
    context = {
        'departments': departments,
        'testimonials': testimonials,
        'total_doctors': Doctor.objects.count(),
        'total_patients': Patient.objects.count(),
    }
    
    return render(request, 'core/index.html', context)


@login_required
def admin_dashboard(request):
    if request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('index')
    
    total_patients = Patient.objects.count()
    total_doctors = Doctor.objects.count()
    total_appointments = Appointment.objects.count()
    pending_appointments = Appointment.objects.filter(status='pending').count()
    
    recent_appointments = Appointment.objects.all().order_by('-created_at')[:10]
    
    # Department-wise statistics
    department_stats = Doctor.objects.values('specialization').annotate(count=Count('id'))
    
    context = {
        'total_patients': total_patients,
        'total_doctors': total_doctors,
        'total_appointments': total_appointments,
        'pending_appointments': pending_appointments,
        'recent_appointments': recent_appointments,
        'department_stats': department_stats,
    }
    
    return render(request, 'core/admin_dashboard.html', context)


@login_required
def manage_doctors(request):
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('index')
    
    doctors = Doctor.objects.all().select_related('user')
    return render(request, 'core/manage_doctors.html', {'doctors': doctors})


@login_required
def manage_patients(request):
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('index')
    
    patients = Patient.objects.all().select_related('user')
    return render(request, 'core/manage_patients.html', {'patients': patients})


@login_required
def analytics(request):
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('index')
    
    from core.ai_utils import doctor_allocator
    
    # Monthly appointment statistics
    from django.db.models.functions import TruncMonth
    monthly_appointments = Appointment.objects.annotate(
        month=TruncMonth('appointment_date')
    ).values('month').annotate(count=Count('id')).order_by('month')
    
    # Department-wise patient distribution
    department_patients = Appointment.objects.values(
        'doctor__specialization'
    ).annotate(count=Count('patient', distinct=True))
    
    # AI-powered workload analytics
    workload_analytics = doctor_allocator.get_workload_analytics()
    
    context = {
        'monthly_appointments': list(monthly_appointments),
        'department_patients': list(department_patients),
        'workload_analytics': workload_analytics,
    }
    
    return render(request, 'core/analytics.html', context)


@login_required
def export_analytics_csv(request):
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('index')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="analytics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow(['Doctor Name', 'Specialization', 'Upcoming Appointments', 'Workload Status', 'Utilization Rate'])
    
    # Get analytics data
    from core.ai_utils import doctor_allocator
    analytics = doctor_allocator.get_workload_analytics()
    
    # Write data rows
    for item in analytics:
        writer.writerow([
            item['doctor'].user.get_full_name(),
            item['doctor'].get_specialization_display(),
            item['workload'],
            item['status'],
            f"{item['utilization_rate']:.1f}%"
        ])
    
    return response


@login_required
def export_analytics_excel(request):
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('index')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analytics Report"
    
    # Get analytics data
    from core.ai_utils import doctor_allocator
    analytics = doctor_allocator.get_workload_analytics()
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Write headers
    headers = ['Doctor Name', 'Specialization', 'Upcoming Appointments', 'Workload Status', 'Utilization Rate']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    for row, item in enumerate(analytics, 2):
        data = [
            item['doctor'].user.get_full_name(),
            item['doctor'].get_specialization_display(),
            item['workload'],
            item['status'],
            f"{item['utilization_rate']:.1f}%"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Color code workload status
            if col == 4:  # Status column
                if item['status'] == 'Low':
                    cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                elif item['status'] == 'Moderate':
                    cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                else:  # High
                    cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    
    # Adjust column widths
    column_widths = [25, 15, 20, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Save to memory
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="analytics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@login_required
def export_analytics_pdf(request):
    if request.user.role != 'admin':
        messages.error(request, 'Access denied')
        return redirect('index')
    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Get analytics data
    from core.ai_utils import doctor_allocator
    analytics = doctor_allocator.get_workload_analytics()
    
    # Create content
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = styles['Heading1']
    title_style.alignment = 1  # Center alignment
    elements.append(Paragraph("Analytics & Reports", title_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    elements.append(Paragraph("<br/><br/>", styles['Normal']))
    
    # Table data
    table_data = [['Doctor Name', 'Specialization', 'Appointments', 'Status', 'Utilization']]
    
    for item in analytics:
        table_data.append([
            item['doctor'].user.get_full_name(),
            item['doctor'].get_specialization_display(),
            str(item['workload']),
            item['status'],
            f"{item['utilization_rate']:.1f}%"
        ])
    
    # Create table
    table = Table(table_data, repeatRows=1)
    
    # Style table
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#DEE2E6')),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ])
    
    table.setStyle(style)
    
    # Add color coding for status rows
    for i, item in enumerate(analytics, 1):
        if item['status'] == 'Low':
            style.add('BACKGROUND', (3, i), (3, i), colors.HexColor('#D4EDDA'))
        elif item['status'] == 'Moderate':
            style.add('BACKGROUND', (3, i), (3, i), colors.HexColor('#FFF3CD'))
        else:  # High
            style.add('BACKGROUND', (3, i), (3, i), colors.HexColor('#F8D7DA'))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="analytics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response


def about(request):
    return render(request, 'core/about.html')


def contact(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        message = request.POST.get('message')
        
        # Here you would typically send an email or save to database
        messages.success(request, 'Thank you for contacting us! We will get back to you soon.')
        return redirect('contact')
    
    return render(request, 'core/contact.html')
